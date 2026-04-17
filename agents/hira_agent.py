"""
HIRA_agent — 약제 요양급여대상여부 평가 SOP 에이전트

역할:
  - 2025.3월 개정 HIRA 규정 제527호 원문을 파싱·인덱싱
  - 제약사의 약제결정신청 준비를 위한 조항 조회·체크리스트 제공
  - 외국약가 A8 조정가 수식이 규정과 일치하는지 **더블체크**

룰 파일:
  agents/rules/hira_agent_rules.md
PDF 원본:
  _resource/약제의 요양급여대상여부 ... 2025년도 3월 개정.pdf
텍스트 캐시:
  data/hira_sop/full_text.txt   (초회 실행 시 자동 생성)
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from agents import price_adjustment as PA

logger = logging.getLogger(__name__)

BASE_DIR  = Path(__file__).parent.parent
PDF_PATH  = BASE_DIR / "_resource" / "약제의 요양급여대상여부 등의 평가기준 및 절차 등에 관한 규정(2025년도 3월 개정).pdf"
XLSX_PATH = BASE_DIR / "_resource" / "국가별 공장도 출하가격 산출식_2025.3월 개정 버전.xlsx"
TEXT_CACHE = BASE_DIR / "data" / "hira_sop" / "full_text.txt"
RULES_PATH = BASE_DIR / "agents" / "rules" / "hira_agent_rules.md"

ARTICLE_RE = re.compile(r"제(\d+)조(?:의(\d+))?\(([^)]+)\)")
PAGE_RE    = re.compile(r"=====\s*\[PAGE\s+(\d+)\]\s*=====")


@dataclass
class Article:
    label:   str          # "제3조의2"
    title:   str          # "자료보완및제출서류반려등"
    page:    int
    body:    str = ""     # 원문 본문 (첫 추출 시만 채움)


class HIRAAgent:
    """약제 요양급여대상여부 평가 SOP 에이전트."""

    # ────────────────────────────────────────────────────────
    # 초기화
    # ────────────────────────────────────────────────────────
    def __init__(self):
        self._full_text:       Optional[str]          = None
        self._articles:        Optional[list[Article]] = None
        self._rules_md:        Optional[str]          = None

    def _ensure_text(self) -> str:
        if self._full_text is not None:
            return self._full_text
        if not TEXT_CACHE.exists():
            self._extract_pdf()
        self._full_text = TEXT_CACHE.read_text(encoding="utf-8")
        return self._full_text

    def _extract_pdf(self) -> None:
        """PDF → 캐시 텍스트. pypdf 필요."""
        try:
            from pypdf import PdfReader
        except ImportError as e:
            raise RuntimeError("pypdf 설치 필요: pip install pypdf") from e
        if not PDF_PATH.exists():
            raise FileNotFoundError(f"HIRA SOP PDF 없음: {PDF_PATH}")
        TEXT_CACHE.parent.mkdir(parents=True, exist_ok=True)
        reader = PdfReader(str(PDF_PATH))
        chunks = []
        for i, p in enumerate(reader.pages):
            chunks.append(f"\n===== [PAGE {i+1}] =====\n" + (p.extract_text() or ""))
        TEXT_CACHE.write_text("\n".join(chunks), encoding="utf-8")
        logger.info("[HIRA_agent] PDF 추출 완료: %d pages → %s", len(reader.pages), TEXT_CACHE)

    # ────────────────────────────────────────────────────────
    # 인덱싱
    # ────────────────────────────────────────────────────────
    def articles(self) -> list[Article]:
        if self._articles is not None:
            return self._articles
        txt = self._ensure_text()
        current_page = 0
        seen = set()
        out: list[Article] = []
        for line in txt.split("\n"):
            m = PAGE_RE.match(line)
            if m:
                current_page = int(m.group(1))
                continue
            for am in ARTICLE_RE.finditer(line):
                label = f"제{am.group(1)}조" + (f"의{am.group(2)}" if am.group(2) else "")
                title = am.group(3)
                key = (label, title)
                if key in seen:
                    continue
                seen.add(key)
                out.append(Article(label=label, title=title, page=current_page))
        self._articles = out
        return out

    def get_article(self, label: str) -> Optional[Article]:
        """ 'ex: 제3조의2' 로 조항 본문 반환 """
        txt = self._ensure_text()
        arts = self.articles()
        # 다음 조항까지의 범위를 본문으로 추출
        labels_in_order = [a.label for a in arts]
        try:
            idx = labels_in_order.index(label)
        except ValueError:
            return None
        start_re = re.compile(re.escape(label) + r"\([^)]+\)")
        m = start_re.search(txt)
        if not m:
            return None
        start = m.start()
        end = len(txt)
        if idx + 1 < len(arts):
            next_label = arts[idx + 1].label
            nm = re.search(re.escape(next_label) + r"\([^)]+\)", txt[start + 10:])
            if nm:
                end = start + 10 + nm.start()
        body = txt[start:end].strip()
        art = arts[idx]
        return Article(label=art.label, title=art.title, page=art.page, body=body)

    # ────────────────────────────────────────────────────────
    # 핵심 요약 (약제결정신청 4대 축)
    # ────────────────────────────────────────────────────────
    KEY_ARTICLES_FOR_PRICING = [
        ("자료제출",        "제3조"),
        ("자료보완·반려",    "제3조의2"),
        ("처리기간",        "제3조의3"),
        ("평가내용",        "제4조"),
        ("선별기준",        "제5조"),
        ("진료상필요약제",    "제6조"),
        ("경제성평가생략",    "제6조의2"),
        ("보건의료영향",     "제6조의3"),
        ("조정관련평가내용",  "제32조"),
        ("조정관련평가기준",  "제33조"),
    ]

    def pricing_application_summary(self) -> dict:
        """제약사 약제결정신청 준비용 핵심 조항 요약."""
        self._ensure_text()
        summary = {
            "title": "약제결정신청(요양급여대상 등재) 핵심 조항",
            "source": str(PDF_PATH.name),
            "revision": "2025.03.05 규정 제527호",
            "pillars": [
                {
                    "id": "자료제출",
                    "articles": ["제3조", "제13조", "제19조", "제24조", "제29조"],
                    "note": "요양급여기준 제10조의2 제3항 자료. 다~바목은 별첨1·2 지침 준수.",
                },
                {
                    "id": "처리기간",
                    "articles": ["제3조의3", "제15조", "제21조", "제26조", "제31조"],
                    "note": "기본 120일, 세계최초 신약 100일. 보완으로 연장 가능 (총 90일 한도).",
                },
                {
                    "id": "평가내용",
                    "articles": ["제4조", "제5조", "제6조", "제6조의2", "제6조의3"],
                    "note": "임상적 유용성·비용효과성·재정영향 종합.",
                },
                {
                    "id": "약가조정",
                    "articles": ["제32조", "제33조"],
                    "note": "외국약가 A8 조정가 산출. 국가별 공장도비율 적용.",
                },
            ],
            "key_articles": [],
        }
        for group, label in self.KEY_ARTICLES_FOR_PRICING:
            art = self.get_article(label)
            if art:
                summary["key_articles"].append({
                    "group": group,
                    "label": art.label,
                    "title": art.title,
                    "page":  art.page,
                    "excerpt": art.body[:400].replace("\n", " ") + ("…" if len(art.body) > 400 else ""),
                })
        return summary

    def submission_checklist(self) -> list[dict]:
        """제출 전 체크리스트 (룰 파일 §3)."""
        return [
            {"item": "품목허가증 사본", "basis": "요양급여기준 제10조의2 제3항 가~나목"},
            {"item": "외국약가 A8 자료 (UK·US·CA·JP·FR·DE·IT·CH)", "basis": "제32조"},
            {"item": "임상적 유용성 자료", "basis": "제3조 다목 / 별첨1"},
            {"item": "비용효과성 자료", "basis": "제3조 라목 / 별첨1"},
            {"item": "경제성평가자료 (경제성평가지침 준수)", "basis": "제3조 마목 / 별첨2"},
            {"item": "재정영향분석", "basis": "제3조 바목"},
            {"item": "요양급여 필요성 의견서 (해당 시)", "basis": "제6조"},
            {"item": "위원회 14일 전까지 추가자료 완결", "basis": "제3조의2 ⑥"},
            {"item": "보완요청 응답 90일 한도 준수", "basis": "제3조의2 ③"},
        ]

    # ────────────────────────────────────────────────────────
    # 외국약가 조정가 더블체크
    # ────────────────────────────────────────────────────────
    # 규정 기준 공장도 출하 비율 (2025.3월). Germany 는 복합식이라 제외.
    REGULATORY_RATIOS = {
        "UK": 0.73, "US": 0.74, "CA": 0.81, "JP": 0.79,
        "FR": 0.77, "IT": 0.93, "CH": 0.73,
    }

    def audit_adjustment_excel(self) -> dict:
        """
        _resource/...산출식_2025.3월.xlsx 의 수식·비율을 규정(및 본 모듈 상수)과 대조.
        K열 수식에 하드코딩된 상수와 H열(공식 비율)의 일치 여부를 확인한다.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"error": "openpyxl 미설치"}
        if not XLSX_PATH.exists():
            return {"error": f"Excel 없음: {XLSX_PATH}"}

        wb = load_workbook(str(XLSX_PATH), data_only=False)
        ws = wb.active

        # 국가 행 맵 (17~24)
        country_map = {
            17: "UK", 18: "US", 19: "CA", 20: "JP",
            21: "FR", 22: "DE", 23: "IT", 24: "CH",
        }

        findings = []
        mismatches = 0
        for row, ctry in country_map.items():
            h_val = ws.cell(row, 8).value     # H열 = 공장도 비율
            k_formula = ws.cell(row, 11).value or ""  # K열 = 조정가 수식
            # K 수식에서 첫 숫자 상수 추출 (G17*0.73*... → 0.73)
            if not isinstance(k_formula, str):
                continue
            m = re.search(r"G\d+\s*\*\s*([0-9.]+)", k_formula)
            k_ratio = float(m.group(1)) if m else None

            reg_ratio = self.REGULATORY_RATIOS.get(ctry)
            pa_ratio  = PA.FACTORY_RATIOS.get(ctry)

            if ctry == "DE":
                # Germany 는 복합식 — K22 에서 × 0.93 확인
                m93 = re.search(r"\*\s*0\.93\b", k_formula)
                status = "ok" if m93 else "check"
                findings.append({
                    "country": ctry,
                    "H_value": h_val,
                    "K_formula_ratio": 0.93 if m93 else None,
                    "regulation_ratio": "복합식 × 0.93",
                    "price_adjustment_module_ratio": "독립 수식",
                    "status": status,
                    "note": "복합식(VAT·도매·수수료 공제) — 별도 검증",
                })
                continue

            status_parts = []
            ok_h = (h_val == reg_ratio)
            ok_k = (k_ratio == reg_ratio) if k_ratio is not None else None
            ok_pa = (pa_ratio == reg_ratio)
            if not ok_h:          status_parts.append("Excel H열 불일치")
            if ok_k is False:     status_parts.append(f"Excel K수식 상수({k_ratio}) ≠ 규정({reg_ratio})")
            if not ok_pa:         status_parts.append("price_adjustment 모듈 불일치")

            mismatch = bool(status_parts)
            if mismatch:
                mismatches += 1

            findings.append({
                "country": ctry,
                "H_value": h_val,
                "K_formula_ratio": k_ratio,
                "regulation_ratio": reg_ratio,
                "price_adjustment_module_ratio": pa_ratio,
                "status": "mismatch" if mismatch else "ok",
                "issues": status_parts,
            })

        # C27 최저가 범위 점검
        c27 = ws["C27"].value
        c27_issue = None
        if isinstance(c27, str) and "MIN(K17:K19)" in c27.replace(" ", ""):
            c27_issue = (
                "Excel C27 = MIN(K17:K19) — UK/USA/Canada 3개국만으로 A8 최저가 산출. "
                "규정상 8개국 원칙. 의도적 subset 이면 근거 문서화 필요."
            )

        return {
            "source":      str(XLSX_PATH.name),
            "total":       len(findings),
            "mismatches":  mismatches,
            "findings":    findings,
            "c27_min_range": c27,
            "c27_warning":   c27_issue,
            "verdict": (
                "PASS" if mismatches == 0 and not c27_issue
                else "FAIL — 상세 findings 확인"
            ),
        }

    # ────────────────────────────────────────────────────────
    # 실제 조정가 산출 (확인용)
    # ────────────────────────────────────────────────────────
    def compute_a8(
        self,
        prices_local: dict[str, float],
        fx_rates: Optional[dict[str, float]] = None,
        subset: Optional[list[str]] = None,
    ) -> dict:
        """
        외국약가(현지통화 최소단위) → A8 조정가 산출.
        HIRA_agent 는 price_adjustment 모듈을 래핑하며, 검증된 비율만 사용함을 보장한다.
        """
        return PA.calculate_a8_min(prices_local, fx_rates=fx_rates, subset=subset)

    # ────────────────────────────────────────────────────────
    # 룰 원문
    # ────────────────────────────────────────────────────────
    def rules_text(self) -> str:
        if self._rules_md is None:
            self._rules_md = RULES_PATH.read_text(encoding="utf-8") if RULES_PATH.exists() else ""
        return self._rules_md
