"""Workbook Exporter — 세션 데이터 → 9-sheet xlsx 파일

입력: 세션 dict (project 메타, prices, scenarios 결과, assumptions, audit log)
출력: openpyxl Workbook 객체 (또는 저장 경로)

호출 시 generate_workbench_sample.py 의 로직을 재사용. 일부 시트는 Phase 1 에서 빈 템플릿.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 스타일
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="1F4E78", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

COUNTRY_NAME = {
    "JP": "일본", "IT": "이탈리아", "FR": "프랑스",
    "CH": "스위스", "UK": "영국", "DE": "독일", "US": "미국",
}


def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _style_subheader(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_cover(wb: Workbook, project: dict, selected: dict):
    ws = wb.create_sheet("Cover")
    _set_widths(ws, [22, 60])
    rows = [
        ("프로젝트", project.get("project_name", "")),
        ("제품명 (EN)", project.get("drug_name_en", "")),
        ("제품명 (KR)", project.get("drug_name_kr", "")),
        ("제조사", project.get("manufacturer", "")),
        ("ATC", project.get("atc", "")),
        ("국내 기준 SKU", project.get("sku", "")),
        ("협상 유형", project.get("neg_type", "")),
        ("", ""),
        ("선정 시나리오", selected.get("name", "")),
        ("제안 상한가 (KRW)", f"₩ {selected.get('proposed_ceiling', 0):,}" if selected.get("proposed_ceiling") else "—"),
        ("산정 기준", selected.get("basis", "")),
        ("", ""),
        ("작성자", project.get("author", "")),
        ("작성일", project.get("date", datetime.today().strftime("%Y-%m-%d"))),
        ("버전", project.get("version", "v1.0")),
    ]
    ws["A1"] = "MA Negotiation Workbench — Cover"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells("A1:B1")
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)


def _sheet_a8_summary(wb: Workbook, scenario: dict):
    ws = wb.create_sheet("A8 Summary")
    rows_map = scenario.get("rows", {})
    has_dose = any(r.get("mg_pack_total") is not None for r in rows_map.values())
    ref_mg = scenario.get("reference_mg")

    if has_dose:
        _set_widths(ws, [10, 12, 14, 12, 14, 12, 14, 16, 16, 10])
        headers = ["국가", "원통화", "원시약가", "SKU mg", "per-mg",
                   f"환산약가({ref_mg}mg)" if ref_mg else "환산약가",
                   "환율", "KRW 환산", "조정가(KRW)", "포함"]
    else:
        _set_widths(ws, [10, 12, 14, 12, 14, 16, 16, 12])
        headers = ["국가", "원통화", "현지약가", "환율", "KRW 환산", "공장도가(KRW)", "조정가(KRW)", "포함"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for country, r in rows_map.items():
        ws.cell(row=row, column=1, value=country)
        ws.cell(row=row, column=2, value="")
        if has_dose:
            ws.cell(row=row, column=3, value=r.get("raw_local_price", r.get("local_price")))
            ws.cell(row=row, column=4, value=r.get("mg_pack_total"))
            per_mg = r.get("price_per_mg")
            ws.cell(row=row, column=5, value=round(per_mg, 4) if per_mg is not None else None)
            ws.cell(row=row, column=6, value=r.get("local_price"))
            ws.cell(row=row, column=7, value=r.get("fx_rate"))
            ws.cell(row=row, column=8, value=r.get("krw_converted"))
            ws.cell(row=row, column=9, value=r.get("adjusted"))
            ws.cell(row=row, column=10, value=r.get("dose_confidence", "포함"))
        else:
            ws.cell(row=row, column=3, value=r.get("local_price"))
            ws.cell(row=row, column=4, value=r.get("fx_rate"))
            ws.cell(row=row, column=5, value=r.get("krw_converted"))
            ws.cell(row=row, column=6, value=r.get("factory_krw"))
            ws.cell(row=row, column=7, value=r.get("adjusted"))
            ws.cell(row=row, column=8, value="포함")
        row += 1

    # 제외 국가
    excluded = scenario.get("excluded") or {}
    if excluded:
        row += 1
        ws.cell(row=row, column=1, value="제외 국가").font = Font(bold=True, color="B91C1C")
        row += 1
        for country, reason in excluded.items():
            ws.cell(row=row, column=1, value=country)
            ws.cell(row=row, column=2, value=reason)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))
            row += 1

    # 통계
    stats = scenario.get("stats", {})
    adjusted_col = 9 if has_dose else 7
    row += 1
    ws.cell(row=row, column=1, value="통계").font = Font(bold=True)
    ws.cell(row=row+1, column=1, value="A8 최저")
    ws.cell(row=row+1, column=adjusted_col, value=stats.get("min"))
    ws.cell(row=row+2, column=1, value="A8 평균")
    ws.cell(row=row+2, column=adjusted_col, value=stats.get("avg"))
    pct = int(stats.get("percent", 0) * 100)
    ws.cell(row=row+3, column=1, value=f"최저×{pct}%")
    ws.cell(row=row+3, column=adjusted_col, value=stats.get("min_percent"))
    ws.cell(row=row+4, column=1, value=f"평균×{pct}%")
    ws.cell(row=row+4, column=adjusted_col, value=stats.get("avg_percent"))
    if has_dose and ref_mg:
        ws.cell(row=row+5, column=1, value="기준 mg").font = Font(italic=True)
        ws.cell(row=row+5, column=2, value=f"{ref_mg} mg — 각국 가격이 이 용량으로 환산됨")


def _sheet_adjustment_logic(wb: Workbook, scenario: dict):
    ws = wb.create_sheet("Adjustment Logic")
    rows_map = scenario.get("rows", {})
    has_dose = any(r.get("mg_pack_total") is not None for r in rows_map.values())
    ref_mg = scenario.get("reference_mg")

    if has_dose:
        _set_widths(ws, [8, 10, 12, 10, 10, 12, 12, 14, 10, 12, 10, 12, 14, 8, 14, 10, 14])
        equiv_label = f"환산약가({ref_mg}mg)" if ref_mg else "환산약가"
        headers = [
            "국가", "자료원", "원시약가", "SKU mg", "per-mg", equiv_label, "dose등급",
            "환율", "KRW환산", "공장도%", "공장도(현지)", "공장도(KRW)",
            "VAT%", "VAT적용", "유통마진%", "조정가(KRW)", "비고",
        ]
    else:
        _set_widths(ws, [8, 12, 12, 10, 12, 10, 12, 14, 8, 14, 10, 14])
        headers = ["국가", "자료원", "현지약가", "환율", "KRW환산", "공장도%",
                   "공장도(현지)", "공장도(KRW)", "VAT%", "VAT적용", "유통마진%", "조정가(KRW)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    for country, r in rows_map.items():
        ws.cell(row=row, column=1, value=country)
        ws.cell(row=row, column=2, value="scraper")
        if has_dose:
            per_mg = r.get("price_per_mg")
            ws.cell(row=row, column=3, value=r.get("raw_local_price", r.get("local_price")))
            ws.cell(row=row, column=4, value=r.get("mg_pack_total"))
            ws.cell(row=row, column=5, value=round(per_mg, 4) if per_mg is not None else None)
            ws.cell(row=row, column=6, value=r.get("local_price"))
            ws.cell(row=row, column=7, value=r.get("dose_confidence", "—"))
            ws.cell(row=row, column=8, value=r.get("fx_rate"))
            ws.cell(row=row, column=9, value=r.get("krw_converted"))
            ws.cell(row=row, column=10, value=r.get("factory_ratio"))
            ws.cell(row=row, column=11, value=r.get("factory_local"))
            ws.cell(row=row, column=12, value=r.get("factory_krw"))
            ws.cell(row=row, column=13, value=r.get("vat_rate"))
            ws.cell(row=row, column=14, value=r.get("vat_applied"))
            ws.cell(row=row, column=15, value=r.get("margin_rate"))
            ws.cell(row=row, column=16, value=r.get("adjusted")).font = Font(bold=True)
            ws.cell(row=row, column=17, value=r.get("form") or "")
        else:
            ws.cell(row=row, column=3, value=r["local_price"])
            ws.cell(row=row, column=4, value=r["fx_rate"])
            ws.cell(row=row, column=5, value=r["krw_converted"])
            ws.cell(row=row, column=6, value=r["factory_ratio"])
            ws.cell(row=row, column=7, value=r["factory_local"])
            ws.cell(row=row, column=8, value=r["factory_krw"])
            ws.cell(row=row, column=9, value=r["vat_rate"])
            ws.cell(row=row, column=10, value=r["vat_applied"])
            ws.cell(row=row, column=11, value=r["margin_rate"])
            ws.cell(row=row, column=12, value=r["adjusted"]).font = Font(bold=True)
        row += 1

    # dose 설명 블록 (normalize 적용 시)
    if has_dose:
        row += 1
        ws.cell(row=row, column=1, value="동등비교 로직").font = Font(bold=True, color="1F4E78")
        row += 1
        ws.cell(row=row, column=1, value=(
            f"각국 SKU의 pack·strength 를 mg 로 파싱한 뒤 per-mg 단가를 산출하고, "
            f"기준 {ref_mg}mg 로 환산한 값(='환산약가')을 공장도% 이후 조정식에 투입합니다. "
            f"dose등급: parsed=파싱, reference=표준SKU 폴백, combo=복합제(제외), unknown=파싱실패."
        ))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        ws.cell(row=row, column=1).alignment = LEFT
        ws.row_dimensions[row].height = 44

    # 제외 국가
    excluded = scenario.get("excluded") or {}
    if excluded:
        row += 1
        ws.cell(row=row, column=1, value="제외 국가").font = Font(bold=True, color="B91C1C")
        row += 1
        for country, reason in excluded.items():
            ws.cell(row=row, column=1, value=country)
            ws.cell(row=row, column=2, value=reason)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))
            row += 1


def _sheet_source_raw(wb: Workbook, source_raw: list[dict]):
    ws = wb.create_sheet("Source Raw")
    _set_widths(ws, [8, 18, 40, 18, 20, 18, 14, 8, 14])
    headers = ["국가", "사이트", "URL", "조회일시", "검색어", "매칭 제품 ID", "원본 가격", "통화", "비고"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    for i, entry in enumerate(source_raw, start=2):
        for c, key in enumerate(["country","site","url","fetched_at","query","product_id","raw_price","currency","note"], start=1):
            ws.cell(row=i, column=c, value=entry.get(key, ""))


def _sheet_product_matching(wb: Workbook, matching: list[dict]):
    ws = wb.create_sheet("Product Matching")
    _set_widths(ws, [8, 14, 24, 14, 12, 10, 20, 14])
    headers = ["국가", "소스", "추출 제품명", "제형", "강도", "Pack", "국내기준 매칭", "일관성 등급"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    for i, entry in enumerate(matching, start=2):
        for c, key in enumerate(["country","source","product_name","form","strength","pack","kr_reference","grade"], start=1):
            ws.cell(row=i, column=c, value=entry.get(key, ""))


# ── HTA 교차검증 상태 스타일
_HTA_STATUS_META = {
    "agree":     ("✅", "C8E6C9", "합의"),      # 연녹색
    "single":    ("⚠️", "FFE0B2", "단일 소스"), # 연주황
    "narrative": ("📝", "E1BEE7", "서술형"),    # 라벤더
    "conflict":  ("❌", "FFCDD2", "충돌"),       # 연빨강
    "missing":   ("∅", "ECEFF1", "미수집"),    # 회색
}


def _hta_truncate(v, max_len=120):
    """xlsx 셀에 너무 긴 서술을 축약."""
    if v is None:
        return ""
    s = str(v)
    return s if len(s) <= max_len else s[: max_len - 1] + "…"


def _sheet_hta_matrix(wb: Workbook, hta_data):
    """
    Phase 2: Tier-3 다중-LLM HTA 교차검증 결과 렌더링.

    hta_data 구조 (tier3_multi_hta_keytruda.json 과 동일 shape):
      {
        "nice": { "agency": {...}, "consensus": {...}, "matrix": {...}, "flags": [...], "summary": {...} },
        "pbac": { ... },
        "has":  { ... },
        "gba":  { ... },
      }
    또는 빈 값/None 이면 플레이스홀더 유지.
    """
    ws = wb.create_sheet("HTA Matrix")

    # ── 빈 데이터: Phase 1 호환 플레이스홀더
    if not hta_data or not isinstance(hta_data, dict):
        _set_widths(ws, [20, 50])
        ws["A1"] = "HTA Matrix — Phase 2 에서 활성화"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
        ws["A2"] = "Tier-3 AI 다중 LLM 교차검증 결과가 로드되면 이 시트가 4개 기관 × 필드 매트릭스로 채워집니다."
        ws.merge_cells("A2:B2")
        return

    # ── 타이틀
    _set_widths(ws, [20, 14, 40, 34, 34, 34, 16])  # 필드 / 상태 / 합의값 / Gemini / Perplexity / OpenAI / 소스
    ws["A1"] = "HTA Matrix — Tier-3 AI 다중 LLM 교차검증"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "3개 LLM (Gemini 2.5-pro grounded, Perplexity sonar-pro, OpenAI GPT-5) 이 독립적으로 응답 → "
        "필드별 합의 상태를 ✅ 합의 / ⚠ 단일소스 / 📝 서술형 / ❌ 충돌 / ∅ 미수집 로 표시."
    )
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 28

    # ── 기관 요약
    row = 4
    ws.cell(row=row, column=1, value="기관별 요약").font = Font(bold=True, size=11, color="1F4E78")
    row += 1
    sum_headers = ["기관", "국가", "합의 ✅", "충돌 ❌", "단일 ⚠", "서술 📝", "미수집 ∅"]
    for c, h in enumerate(sum_headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, len(sum_headers))
    row += 1
    for code, body in hta_data.items():
        agency = body.get("agency", {})
        s = body.get("summary", {})
        ws.cell(row=row, column=1, value=agency.get("name", code.upper())).font = Font(bold=True)
        ws.cell(row=row, column=2, value=agency.get("country", ""))
        ws.cell(row=row, column=3, value=s.get("agree_count", 0))
        ws.cell(row=row, column=4, value=s.get("conflict_count", 0))
        ws.cell(row=row, column=5, value=s.get("single_source", 0))
        ws.cell(row=row, column=6, value=s.get("narrative_count", 0))
        ws.cell(row=row, column=7, value=s.get("missing_count", 0))
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = CENTER
        row += 1
    row += 2

    # ── 기관별 상세 매트릭스
    for code, body in hta_data.items():
        agency = body.get("agency", {})
        matrix = body.get("matrix", {})
        consensus = body.get("consensus", {})
        flags = body.get("flags", [])
        fields = agency.get("fields", list(consensus.keys()))

        # 기관 타이틀
        ws.cell(row=row, column=1, value=f"{agency.get('name', code.upper())} ({agency.get('country', '')})").font = Font(bold=True, size=12, color="1F4E78")
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        # 매트릭스 헤더
        headers = ["필드", "상태", "합의 값", "Gemini", "Perplexity", "OpenAI", "소스 수"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1

        # 필드 행
        for field in fields:
            cons = consensus.get(field, {})
            status = cons.get("status", "missing")
            icon, fg, label = _HTA_STATUS_META.get(status, ("?", "FFFFFF", status))

            ws.cell(row=row, column=1, value=field).font = Font(bold=True)

            # 상태 셀
            status_cell = ws.cell(row=row, column=2, value=f"{icon} {label}")
            status_cell.fill = PatternFill("solid", fgColor=fg)
            status_cell.alignment = CENTER

            # 합의 값
            if status == "agree":
                cons_val = _hta_truncate(cons.get("value"))
            elif status == "single":
                cons_val = f"[{cons.get('sources', [''])[0]}] {_hta_truncate(cons.get('value'))}"
            elif status == "narrative":
                cons_val = "서술형 — 소스별 표현 차이, 통합 필요"
            elif status == "conflict":
                cons_val = "⚠ 충돌: " + " vs ".join(_hta_truncate(v, 40) for v in (cons.get("values") or {}).values() if v is not None)
            else:
                cons_val = "—"
            ws.cell(row=row, column=3, value=cons_val).alignment = LEFT

            # LLM별 값
            field_row = matrix.get(field, {})
            ws.cell(row=row, column=4, value=_hta_truncate(field_row.get("gemini"))).alignment = LEFT
            ws.cell(row=row, column=5, value=_hta_truncate(field_row.get("perplexity"))).alignment = LEFT
            ws.cell(row=row, column=6, value=_hta_truncate(field_row.get("openai"))).alignment = LEFT
            ws.cell(row=row, column=7, value=len(cons.get("sources", []))).alignment = CENTER

            for c in range(1, 8):
                ws.cell(row=row, column=c).border = BORDER
            ws.row_dimensions[row].height = 36
            row += 1

        # 플래그
        if flags:
            ws.cell(row=row, column=1, value="플래그").font = Font(bold=True, italic=True, color="B45309")
            for c in range(2, 8):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFF8EB")
            row += 1
            for f in flags:
                sev = f.get("severity") or "warn"
                prefix = "ⓘ" if sev == "info" else "⚠"
                ws.cell(row=row, column=1, value=f"{prefix} {f.get('field', '')}")
                ws.cell(row=row, column=2, value=f.get("issue", "")).alignment = LEFT
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
                ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="FFF8EB")
                row += 1

        row += 2  # 기관 간 간격


def _sheet_assumptions(wb: Workbook, assumptions: dict):
    ws = wb.create_sheet("Assumptions")
    _set_widths(ws, [24, 40])
    ws["A1"] = "환율 기본값"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "윈도우 (개월)"
    ws["B2"] = assumptions.get("fx_window_months", 36)
    ws["A3"] = "출처"
    ws["B3"] = assumptions.get("fx_source", "")

    ws["A5"] = "국가별 가정치"
    ws["A5"].font = Font(bold=True)
    headers = ["국가", "통화", "공장도%", "VAT%", "유통마진%", "환율 기본값"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    _style_header(ws, 6, len(headers))
    _set_widths(ws, [8, 10, 12, 10, 14, 16])

    for i, (country, data) in enumerate(assumptions.get("countries", {}).items(), start=7):
        ws.cell(row=i, column=1, value=country)
        ws.cell(row=i, column=2, value=data.get("currency", ""))
        ws.cell(row=i, column=3, value=data.get("factory_ratio"))
        ws.cell(row=i, column=4, value=data.get("vat_rate"))
        ws.cell(row=i, column=5, value=data.get("margin_rate"))
        ws.cell(row=i, column=6, value=data.get("fx_rate_default"))


def _sheet_scenarios(wb: Workbook, scenarios: list[dict], selected_name: str):
    ws = wb.create_sheet("Scenarios")
    _set_widths(ws, [20] + [18] * len(scenarios))
    ws.cell(row=1, column=1, value="항목").font = Font(bold=True)
    for j, s in enumerate(scenarios, start=2):
        ws.cell(row=1, column=j, value=s["name"])
        if s["name"] == selected_name:
            ws.cell(row=1, column=j).fill = PatternFill("solid", fgColor="C8E6C9")
    _style_subheader(ws, 1, 1 + len(scenarios))

    def row(label, key, fmt=str):
        ws.cell(row=row.n, column=1, value=label).font = Font(bold=True)
        for j, s in enumerate(scenarios, start=2):
            v = s.get(key) if isinstance(s, dict) else None
            if v is not None and fmt is not str:
                ws.cell(row=row.n, column=j, value=fmt(v))
            else:
                ws.cell(row=row.n, column=j, value=str(v) if v is not None else "")
        row.n += 1
    row.n = 2

    row("포함 국가", None)
    # 다중 국가는 수동 채움
    for j, s in enumerate(scenarios, start=2):
        include = ", ".join(s.get("rows", {}).keys())
        ws.cell(row=2, column=j, value=include)

    row.n = 3
    row("공식", "basis")
    row("A8 최저",
        lambda s: s.get("stats", {}).get("min"), fmt=str)
    # stats 중첩은 수동 처리
    for rlabel, statkey in [
        ("A8 최저", "min"),
        ("A8 평균", "avg"),
        ("최저×N%", "min_percent"),
        ("평균×N%", "avg_percent"),
        ("국가 수", "n_countries"),
    ]:
        ws.cell(row=row.n, column=1, value=rlabel).font = Font(bold=True)
        for j, s in enumerate(scenarios, start=2):
            v = s.get("stats", {}).get(statkey)
            ws.cell(row=row.n, column=j, value=v if v is not None else "")
        row.n += 1

    ws.cell(row=row.n, column=1, value="제안 상한가").font = Font(bold=True)
    for j, s in enumerate(scenarios, start=2):
        ws.cell(row=row.n, column=j, value=s.get("proposed_ceiling"))
    row.n += 1
    ws.cell(row=row.n, column=1, value="근거").font = Font(bold=True)
    for j, s in enumerate(scenarios, start=2):
        ws.cell(row=row.n, column=j, value=s.get("notes", ""))


def _sheet_audit_log(wb: Workbook, audit_log: list[dict]):
    ws = wb.create_sheet("Audit Log")
    _set_widths(ws, [20, 14, 16, 16, 20, 20, 30])
    headers = ["타임스탬프", "사용자", "시트/항목", "변경 필드", "이전 값", "신규 값", "사유"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))
    for i, e in enumerate(audit_log or [], start=2):
        for c, key in enumerate(["timestamp","user","sheet","field","old","new","reason"], start=1):
            ws.cell(row=i, column=c, value=e.get(key, ""))


def export_workbook(
    session: dict,
    out_path: Path | str,
) -> Path:
    """
    session 구조:
      {
        "project":     {project_name, drug_name_en/kr, manufacturer, atc, sku, neg_type, author, date, version},
        "assumptions": {...},
        "prices":      {country: local_price},
        "scenarios":   [{name, proposed_ceiling, basis, rows, stats, notes}],   # compute_scenario 결과 리스트
        "selected":    "B안",
        "source_raw":  [{country, site, url, fetched_at, query, product_id, raw_price, currency, note}],
        "matching":    [{country, source, product_name, form, strength, pack, kr_reference, grade}],
        "hta":         [{agency, country, decision, scope, icer, pas_rsa, rationale, confidence}] | None,
        "audit_log":   [{timestamp, user, sheet, field, old, new, reason}],
      }
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    scenarios = session.get("scenarios", [])
    selected_name = session.get("selected") or (scenarios[0]["name"] if scenarios else "")
    selected = next((s for s in scenarios if s["name"] == selected_name), scenarios[0] if scenarios else {})

    _sheet_cover(wb, session.get("project", {}), selected)
    _sheet_a8_summary(wb, selected)
    _sheet_adjustment_logic(wb, selected)
    _sheet_source_raw(wb, session.get("source_raw", []))
    _sheet_product_matching(wb, session.get("matching", []))
    _sheet_hta_matrix(wb, session.get("hta"))
    _sheet_assumptions(wb, session.get("assumptions", {}))
    _sheet_scenarios(wb, scenarios, selected_name)
    _sheet_audit_log(wb, session.get("audit_log", []))

    wb.save(out_path)
    return out_path
