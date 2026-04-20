"""IQVIA NSA-E Master quarterly Excel → market_share_* 테이블 ingester.

입력 시트 구조 (NSA sheet):
  row 0: 그룹 헤더 (quarter 라벨 "QTR ~ 03/2021" 등 — MAT 포함)
  row 1: 필드명 ("OTC/ETHICAL", "ATC 4 CODE", "PRODUCT NAME", ... "Values LC 3/2021", "Dosage Units 3/2021")
  row 2+: 데이터

unique key = hash(PRODUCT NAME + MOLECULE DESC + PACK)
동일 quarter 에 같은 key 중복시 sum.
MAT 컬럼은 스킵 (quarterly 데이터만 적재 — MAT 은 API 에서 계산).
"""
from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl


# ─── 필드명 → 컬럼 매핑 ──────────────────────────────────────────────
META_FIELDS = {
    "OTC/ETHICAL": "otc_ethical",
    "ATC 2 CODE": "atc2_code",
    "ATC 3 CODE": "atc3_code",
    "ATC 4 CODE": "atc4_code",
    "ATC 4 DESC": "atc4_desc",
    "MFR NAME": "mfr_name",
    "PRODUCT NAME": "product_name",
    "PRODUCT GROUP": "product_group",
    "MOLECULE DESC": "molecule_desc",
    "CORP": "corp",
    "13MNC": "mnc13",
    "EM-ethical": "em_ethical",
    "KR market": "kr_market",
    "NFC 3": "nfc3",
    "STRENGTH": "strength",
    "PACK DESC": "pack_desc",
    "PACK": "pack",
    "PACK LAUNCH DATE": "pack_launch_date",
}

QUARTER_RE = re.compile(r"(\d{1,2})/(\d{4})")
# "Values LC\n3/2021" / "Dosage Units\n03/2025" / "Values LC MAT Dec 2024"


def _norm(s) -> str:
    return str(s).strip() if s is not None else ""


def _parse_quarter(header_label: str) -> str | None:
    """ "3/2021" -> "2021Q1". MAT 은 None 반환."""
    if "MAT" in header_label.upper():
        return None
    m = QUARTER_RE.search(header_label)
    if not m:
        return None
    month = int(m.group(1))
    year = int(m.group(2))
    q = (month - 1) // 3 + 1
    if q < 1 or q > 4:
        return None
    return f"{year}Q{q}"


def _product_hash(product_name: str, molecule_desc: str, pack: str) -> str:
    raw = f"{product_name}|{molecule_desc}|{pack}".lower().strip()
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def _build_column_map(field_row: tuple, group_row: tuple) -> dict:
    """
    Returns:
      meta_cols: {db_col_name: excel_col_idx}
      value_cols: {(metric, quarter): excel_col_idx}  metric ∈ {"values_lc","dosage_units"}
    """
    meta_cols: dict[str, int] = {}
    value_cols: dict[tuple[str, str], int] = {}

    for idx, cell in enumerate(field_row):
        label = _norm(cell)
        if not label:
            continue
        # 메타
        if label in META_FIELDS:
            meta_cols[META_FIELDS[label]] = idx
            continue
        # 값 컬럼 ("Values LC\n3/2021" or "Dosage Units\n03/2025")
        clean = label.replace("\n", " ").strip()
        if clean.lower().startswith("values lc"):
            metric = "values_lc"
        elif clean.lower().startswith("dosage units"):
            metric = "dosage_units"
        else:
            continue
        q = _parse_quarter(clean)
        if q:
            value_cols[(metric, q)] = idx

    return {"meta": meta_cols, "values": value_cols}


def _rows_iter(ws) -> Iterable[tuple]:
    return ws.iter_rows(values_only=True)


def ingest(
    xlsx_path: Path,
    db_path: Path,
    sheet_name: str = "NSA",
    uploaded_by: str | None = None,
) -> dict:
    """Excel 파싱 → DB 적재. 반환: {rows_ingested, products, quarters}.

    이미 같은 product_id+quarter 레코드는 REPLACE (최신 업로드 우선).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 {sheet_name!r} 없음. 가능: {wb.sheetnames}")
    ws = wb[sheet_name]

    it = _rows_iter(ws)
    group_row = next(it)      # row 0: quarter 라벨
    field_row = next(it)      # row 1: 필드명
    col_map = _build_column_map(field_row, group_row)
    meta_cols = col_map["meta"]
    value_cols = col_map["values"]

    quarters = sorted({q for (_, q) in value_cols.keys()})
    now_iso = datetime.now(timezone.utc).isoformat()

    products: dict[str, dict] = {}
    quarterly_accum: dict[tuple[str, str], dict[str, float]] = {}
    # quarterly_accum[(product_id, quarter)] = {"values_lc": x, "dosage_units": y}

    row_count = 0
    for row in it:
        product_name = _norm(row[meta_cols.get("product_name", -1)]) \
            if "product_name" in meta_cols else ""
        molecule_desc = _norm(row[meta_cols.get("molecule_desc", -1)]) \
            if "molecule_desc" in meta_cols else ""
        pack = _norm(row[meta_cols.get("pack", -1)]) if "pack" in meta_cols else ""

        if not product_name or not molecule_desc:
            continue
        pid = _product_hash(product_name, molecule_desc, pack)

        # 메타 (첫 번째 관측만 유지; 중복시 기존 값 신뢰)
        if pid not in products:
            meta = {"product_id": pid, "updated_at": now_iso}
            for db_col, idx in meta_cols.items():
                v = row[idx] if idx < len(row) else None
                if db_col == "pack_launch_date" and v is not None:
                    # "4/2015" → keep as string
                    meta[db_col] = _norm(v)
                else:
                    meta[db_col] = _norm(v) if v is not None else None
            products[pid] = meta

        # 분기별 값 누적
        for (metric, quarter), idx in value_cols.items():
            v = row[idx] if idx < len(row) else None
            if v is None:
                continue
            try:
                v_num = float(v)
            except (TypeError, ValueError):
                continue
            key = (pid, quarter)
            if key not in quarterly_accum:
                quarterly_accum[key] = {"values_lc": 0.0, "dosage_units": 0.0}
            quarterly_accum[key][metric] += v_num

        row_count += 1

    # DB 저장
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON;")
    try:
        with conn:
            cur = conn.cursor()
            # 1) products
            product_cols = [
                "product_id","otc_ethical","atc2_code","atc3_code","atc4_code","atc4_desc",
                "mfr_name","corp","mnc13","product_name","product_group","molecule_desc",
                "em_ethical","kr_market","nfc3","strength","pack_desc","pack",
                "pack_launch_date","updated_at",
            ]
            placeholders = ",".join(["?"] * len(product_cols))
            upsert_sql = (
                f"INSERT OR REPLACE INTO market_share_products "
                f"({', '.join(product_cols)}) VALUES ({placeholders})"
            )
            rows = [tuple(p.get(c) for c in product_cols) for p in products.values()]
            cur.executemany(upsert_sql, rows)

            # 2) quarterly
            q_rows = [
                (pid, qtr, vals.get("values_lc"), vals.get("dosage_units"))
                for (pid, qtr), vals in quarterly_accum.items()
            ]
            cur.executemany(
                "INSERT OR REPLACE INTO market_share_quarterly "
                "(product_id, quarter, values_lc, dosage_units) VALUES (?,?,?,?)",
                q_rows,
            )

            # 3) upload_log
            cur.execute(
                "INSERT INTO market_share_upload_log "
                "(uploaded_at, uploaded_by, filename, rows_ingested, quarters_json) "
                "VALUES (?,?,?,?,?)",
                (now_iso, uploaded_by, str(xlsx_path.name), row_count,
                 json.dumps(quarters, ensure_ascii=False)),
            )
    finally:
        conn.close()

    return {
        "rows_ingested": row_count,
        "unique_products": len(products),
        "quarterly_points": len(quarterly_accum),
        "quarters": quarters,
    }


if __name__ == "__main__":
    import sys
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 \
        else Path(__file__).parent.parent.parent / "_resource" / "NSA_E_Master_2025Q4.xlsx"
    db = Path(__file__).parent.parent.parent / "data" / "db" / "drug_prices.db"
    print(f"ingesting {xlsx} → {db}")
    result = ingest(xlsx, db, uploaded_by="cli")
    print(json.dumps(result, ensure_ascii=False, indent=2))
