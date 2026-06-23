"""투약비용비교 → 수식이 살아있는 xlsx 생성.

엑셀에서 환자 파라미터·용량을 바꾸면 BSA/GFR·용량(mg)·치료비가 재계산되도록
수식으로 작성(레지멘DB 계산기 방식). 약가(단위가·함량)는 as-of DB 조회 스냅샷이라
값으로 기입하고, 그 외(1회mg·주기총량·사이클/코스/월/연)는 수식.

시트: 계산기(환자) · 비교(레지멘별 요약) · 레지멘별 1시트씩.
"""
from __future__ import annotations

import io
import re

CALC = "계산기"


def _safe_sheet(name: str, used: set, idx: int) -> str:
    s = re.sub(r"[\[\]:*?/\\]", " ", name or f"레지멘{idx}").strip()[:26] or f"레지멘{idx}"
    base = f"{idx}.{s}"[:31]
    out, n = base, 2
    while out in used:
        out = f"{base[:28]}_{n}"
        n += 1
    used.add(out)
    return out


def _calc_sheet(ws, patient: dict):
    p = {"height": 165, "weight": 62, "age": 60, "sex": "M", "scr": 0.9, **(patient or {})}
    ws["A1"] = "환자 파라미터 (노란칸 입력 → 자동 계산)"
    rows = [("키 (cm)", "height"), ("체중 (kg)", "weight"), ("나이", "age"),
            ("성별 (M/F)", "sex"), ("SCr (mg/dL)", "scr")]
    for i, (label, key) in enumerate(rows, start=2):
        ws.cell(i, 1, label)
        ws.cell(i, 2, p[key])
    ws["A8"] = "BSA Mosteller (m²)"; ws["B8"] = "=SQRT(B2*B3/3600)"
    ws["A9"] = "BSA DuBois (m²)"; ws["B9"] = "=0.007184*B2^0.725*B3^0.425"
    ws["A10"] = "적용 BSA (m²)"; ws["B10"] = "=B8"
    ws["A11"] = "CrCl Cockcroft-Gault"; ws["B11"] = '=((140-B4)*B3*IF(B5="F",0.85,1))/(72*B6)'
    ws["A12"] = "적용 GFR (cap125)"; ws["B12"] = "=MIN(B11,125)"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16


_HDR = ["약제", "용량값", "단위", "투여일", "회수/주기", "주기(일)", "총사이클",
        "1회(mg)", "주기총량(mg)", "단위가(₩)", "함량(mg)", "mg당단가",
        "사이클(₩)", "코스(₩)", "월(₩)", "연(₩)", "가격소스/규격"]


def _regimen_sheet(ws, reg: dict):
    """reg: {name, drugs:[{ingredient, dose_value, unit, dose_days, per_cycle, cycle_days,
            total_cycles, price:{unit_price, content_mg, label, available}}]}."""
    for c, h in enumerate(_HDR, start=1):
        ws.cell(1, c, h)
    drugs = reg.get("drugs") or []
    r0 = 2
    for i, d in enumerate(drugs):
        r = r0 + i
        price = d.get("price") or {}
        is_count = (d.get("unit") or "").strip().lower() in (
            "정", "tab", "tablet", "캡슐", "cap", "capsule", "포", "환", "매", "vial", "바이알", "unit", "단위")
        ws.cell(r, 1, d.get("ingredient"))
        ws.cell(r, 2, d.get("dose_value"))
        ws.cell(r, 3, d.get("unit"))
        ws.cell(r, 4, d.get("dose_days"))
        ws.cell(r, 5, d.get("per_cycle") or 1)
        ws.cell(r, 6, d.get("cycle_days") or 1)
        ws.cell(r, 7, d.get("total_cycles"))
        # 1회(mg) — 단위별 수식 (계산기 BSA/체중/GFR 참조)
        ws.cell(r, 8, (
            f'=IF($C{r}="mg/m2",$B{r}*\'{CALC}\'!$B$10,'
            f'IF($C{r}="mg/m2/day",$B{r}*\'{CALC}\'!$B$10,'
            f'IF($C{r}="g/m2",$B{r}*1000*\'{CALC}\'!$B$10,'
            f'IF($C{r}="mg/kg",$B{r}*\'{CALC}\'!$B$3,'
            f'IF($C{r}="mg/kg/day",$B{r}*\'{CALC}\'!$B$3,'
            f'IF($C{r}="AUC",$B{r}*(\'{CALC}\'!$B$12+25),$B{r}))))))'
        ))
        ws.cell(r, 9, f'=IF(N($B{r})=0,"",$H{r}*$E{r})')                 # 주기총량
        up = price.get("unit_price")
        ws.cell(r, 10, up if up is not None else None)                   # 단위가(값)
        # 함량: count 단위는 1(주기총량=count, mg당단가=단위가)
        ws.cell(r, 11, 1 if is_count else price.get("content_mg"))
        ws.cell(r, 12, f'=IF(N($K{r})=0,"",$J{r}/$K{r})')               # mg당단가
        ws.cell(r, 13, f'=IF(OR($I{r}="",$L{r}=""),"",$I{r}*$L{r})')     # 사이클
        ws.cell(r, 14, f'=IF(OR($M{r}="",N($G{r})=0),"",$M{r}*$G{r})')   # 코스
        ws.cell(r, 15, f'=IF(OR($M{r}="",N($F{r})=0),"",$M{r}*30/$F{r})')   # 월
        ws.cell(r, 16, f'=IF(OR($M{r}="",N($F{r})=0),"",$M{r}*365/$F{r})')  # 연
        ws.cell(r, 17, (("가중평균" if (d.get("price_source") or "") != "domestic" else "브랜드")
                        + (f" · {price.get('label')}" if price.get("label") else "")
                        + ("" if price.get("available", True) else " · 가격없음")))
    last = r0 + len(drugs) - 1 if drugs else r0
    tot = last + 1
    ws.cell(tot, 1, "합계")
    for col in (13, 14, 15, 16):
        L = chr(64 + col)
        ws.cell(tot, col, f'=SUM({L}{r0}:{L}{last})' if drugs else 0)
    for c, w in [(1, 22), (3, 12), (9, 13), (10, 12), (12, 11), (13, 13), (14, 13), (15, 13), (16, 13), (17, 30)]:
        ws.column_dimensions[chr(64 + c)].width = w
    return tot  # 합계 행 번호


def build_regimen_xlsx(date: str, source: str, patient: dict, regimens: list[dict]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    calc = wb.active
    calc.title = CALC
    _calc_sheet(calc, patient)

    cmp_ws = wb.create_sheet("비교")
    used = {CALC, "비교"}
    sheet_refs = []  # (regimen_name, sheet_name, total_row, has_cycle)
    for i, reg in enumerate(regimens, start=1):
        sn = _safe_sheet(reg.get("name") or f"레지멘{i}", used, i)
        ws = wb.create_sheet(sn)
        tot = _regimen_sheet(ws, reg)
        has_cycle = any((d.get("cycle_days") or 1) > 1 for d in (reg.get("drugs") or []))
        sheet_refs.append((reg.get("name") or sn, sn, tot, has_cycle))

    # 비교 시트
    cmp_ws["A1"] = f"레지멘별 비용 비교  (기준일 {date} · 소스 {'주성분 가중평균' if source != 'domestic' else '국내약가'})"
    hdr = ["레지멘", "1사이클", "전체코스", "일", "월", "연"]
    for c, h in enumerate(hdr, start=1):
        cell = cmp_ws.cell(3, c, h)
        cell.font = Font(bold=True)
    for i, (name, sn, tot, has_cycle) in enumerate(sheet_refs):
        r = 4 + i
        cmp_ws.cell(r, 1, name)
        q = f"'{sn}'"
        cmp_ws.cell(r, 2, f"={q}!M{tot}" if has_cycle else "—")        # 1사이클
        cmp_ws.cell(r, 3, f"={q}!N{tot}" if has_cycle else "—")        # 전체코스
        cmp_ws.cell(r, 4, f"={q}!P{tot}/365")                          # 일 = 연/365
        cmp_ws.cell(r, 5, f"={q}!O{tot}")                              # 월
        cmp_ws.cell(r, 6, f"={q}!P{tot}")                              # 연
    for c, w in [(1, 30), (2, 14), (3, 14), (4, 12), (5, 14), (6, 16)]:
        cmp_ws.column_dimensions[chr(64 + c)].width = w
    cmp_ws["A2"] = "※ 1사이클·전체코스는 항암 레지멘(주기 투여)만 · 약가(단위가·함량)는 기준일 DB 스냅샷, 그 외 셀은 수식으로 재계산"

    # 통화 서식
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and any(
                        t in cell.value for t in ("!M", "!N", "!O", "!P", "$L", "$J", "$M")):
                    cell.number_format = "#,##0"

    # 노란 입력칸 강조
    yellow = PatternFill("solid", fgColor="FFF6C0")
    for r in range(2, 7):
        calc.cell(r, 2).fill = yellow

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
