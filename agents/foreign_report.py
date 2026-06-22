"""해외약가 A8 리포트 생성 — xlsx(openpyxl) + hwpx(python-hwpx).

첨부 양식: 국가별 컬럼 × (외국약가/병·/정·적용환율·환산가·조정가·국내신청가비율·출처) 행.
데이터원: ForeignPriceAgent.get_cached_results (단일 진실원). 조정가는 용량 정규화가
(normalized) 우선. 국내 신청가(표시가/실제가)는 다운로드 시 입력폼으로 받음.
"""
from __future__ import annotations

import io
from datetime import date

# 표시 순서 (데이터 있는 국가만 노출)
COUNTRY_ORDER = ["US", "UK", "CA", "DE", "FR", "IT", "CH", "JP"]
COUNTRY_KR = {"US": "미국", "UK": "영국", "CA": "캐나다", "DE": "독일",
              "FR": "프랑스", "IT": "이탈리아", "CH": "스위스", "JP": "일본"}
CUR_SYM = {"USD": "$", "GBP": "£", "CAD": "C$", "EUR": "€", "JPY": "¥", "CHF": "CHF"}


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_report_model(cached: dict, applied_display=None, applied_actual=None) -> dict:
    """get_cached_results 결과 → A8 리포트 모델 (컬럼=국가, 행=지표)."""
    cols = []
    for c in COUNTRY_ORDER:
        rows = cached.get(c) or []
        priced = [r for r in rows if r.get("adjusted_price_krw") is not None]
        if not priced:
            continue
        r = priced[0]
        per_unit = _num(r.get("per_unit_local"))
        fx = _num(r.get("exchange_rate"))
        krw_conv = _num(r.get("krw_converted"))
        if krw_conv is None and per_unit is not None and fx is not None:
            krw_conv = per_unit * fx
        adj = _num(r.get("adjusted_price_krw_normalized")) or _num(r.get("adjusted_price_krw"))
        cols.append({
            "country": c,
            "country_kr": COUNTRY_KR.get(c, c),
            "currency": r.get("currency") or "",
            "sym": CUR_SYM.get(r.get("currency") or "", ""),
            "local_price": _num(r.get("local_price")),
            "per_unit_local": per_unit,
            "pack_count": int(r.get("pack_count") or 1),
            "exchange_rate": fx,
            "fx_from": r.get("exchange_rate_from") or "",
            "fx_to": r.get("exchange_rate_to") or "",
            "krw_converted": krw_conv,
            "adjusted": adj,
            "source": r.get("source_label") or "",
            "dose_norm_note": r.get("dose_norm_note") or "",
        })

    actual = _num(applied_actual)
    adj_vals = [c["adjusted"] for c in cols if c["adjusted"] is not None]
    conv_vals = [c["krw_converted"] for c in cols if c["krw_converted"] is not None]
    avg_adj = sum(adj_vals) / len(adj_vals) if adj_vals else None
    avg_conv = sum(conv_vals) / len(conv_vals) if conv_vals else None
    # 국내신청가비율: 국가별 = 조정가/실제가, 평균칸 = 실제가/평균조정가 (양식 관례)
    for c in cols:
        c["ratio"] = (round(c["adjusted"] / actual * 100) if (actual and c["adjusted"]) else None)
    avg_ratio = round(actual / avg_adj * 100) if (actual and avg_adj) else None

    return {
        "cols": cols,
        "avg": {"krw_converted": avg_conv, "adjusted": avg_adj, "ratio": avg_ratio},
        "applied_display": _num(applied_display),
        "applied_actual": actual,
        "fx_period": (cols[0]["fx_from"], cols[0]["fx_to"]) if cols else ("", ""),
        "pack_count": cols[0]["pack_count"] if cols else 1,
    }


def _rows_for_render(m: dict) -> tuple[list[str], list[list[str]]]:
    """(header, rows) — 첫 열은 행라벨, 이후 국가별, 마지막 평균."""
    cols = m["cols"]
    pack = m["pack_count"]
    header = ["등재국가 (통화)"] + [f"{c['country_kr']} ({c['sym']})" for c in cols] + ["평균"]

    def money(v, dash="-"):
        return f"{round(v):,}" if v is not None else dash

    def local(v, dash="-"):
        return f"{v:,.2f}" if v is not None else dash

    body = [
        [f"외국약가/병 (={pack}정) (해당국 화폐)"] + [local(c["local_price"]) for c in cols] + ["-"],
        ["외국약가/정 (해당국 화폐)"] + [local(c["per_unit_local"]) for c in cols] + ["-"],
        ["적용환율 (36개월 최종 평균 매매기준율)"] + [local(c["exchange_rate"]) for c in cols] + ["-"],
        ["환산가 (원)"] + [money(c["krw_converted"]) for c in cols] + [money(m["avg"]["krw_converted"])],
        ["조정가 (원)"] + [money(c["adjusted"]) for c in cols] + [money(m["avg"]["adjusted"])],
        ["국내신청가격비율 (실제가) (%)"]
            + [f"{c['ratio']}%" if c["ratio"] is not None else "-" for c in cols]
            + [f"{m['avg']['ratio']}%" if m["avg"]["ratio"] is not None else "-"],
        ["근거 (출처)"] + [c["source"] or "-" for c in cols] + ["-"],
    ]
    return header, body


def _title(m: dict, query: str) -> str:
    disp = f"{round(m['applied_display']):,}" if m["applied_display"] else "-"
    act = f"{round(m['applied_actual']):,}" if m["applied_actual"] else "-"
    return f"{query} A8 약가 비교 (국내 신청가격: 표시가 {disp}원/정, 실제가 {act}/정)"


def export_xlsx(query: str, model: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    THIN = Side(border_style="thin", color="BBBBBB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HFILL = PatternFill("solid", fgColor="00857C")
    LBLFILL = PatternFill("solid", fgColor="EAF2F1")

    header, body = _rows_for_render(model)
    wb = Workbook()
    ws = wb.active
    ws.title = "A8 비교"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    tc = ws.cell(row=1, column=1, value=_title(model, query))
    tc.font = Font(bold=True, size=12)
    tc.alignment = CENTER
    # 헤더
    for j, h in enumerate(header, 1):
        cc = ws.cell(row=2, column=j, value=h)
        cc.fill = HFILL
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.alignment = CENTER
        cc.border = BORDER
    # 본문
    for i, row in enumerate(body, 3):
        for j, val in enumerate(row, 1):
            cc = ws.cell(row=i, column=j, value=val)
            cc.alignment = CENTER if j > 1 else Alignment(vertical="center", wrap_text=True)
            cc.border = BORDER
            if j == 1:
                cc.fill = LBLFILL
                cc.font = Font(bold=True, size=10)
    ws.column_dimensions["A"].width = 30
    for j in range(2, len(header) + 1):
        ws.column_dimensions[chr(64 + j) if j <= 26 else "A"].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_hwpx(query: str, model: dict) -> bytes:
    from hwpx.document import HwpxDocument
    header, body = _rows_for_render(model)
    ncols = len(header)
    nrows = len(body) + 1  # +헤더

    doc = HwpxDocument.new()
    doc.add_paragraph(_title(model, query))
    table = doc.add_table(nrows, ncols)
    for j, h in enumerate(header):
        table.set_cell_text(0, j, str(h))
    for i, row in enumerate(body, 1):
        for j, val in enumerate(row):
            table.set_cell_text(i, j, str(val))
    return doc.to_bytes()


def generate(query: str, fmt: str, cached: dict,
             applied_display=None, applied_actual=None) -> tuple[bytes, str, str]:
    """리포트 생성 → (bytes, mimetype, filename)."""
    model = build_report_model(cached, applied_display, applied_actual)
    if not model["cols"]:
        raise ValueError("리포트 생성할 가격 데이터가 없습니다.")
    stamp = date.today().isoformat()
    if fmt == "hwpx":
        return (export_hwpx(query, model),
                "application/octet-stream", f"{query}_A8_report_{stamp}.hwpx")
    return (export_xlsx(query, model),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{query}_A8_report_{stamp}.xlsx")
