"""
MA Negotiation Workbench — xlsx 템플릿 샘플 생성기
- Keytruda (pembrolizumab) 100mg/4mL 1 vial 기준
- 실측 데이터: CH, DE, IT, UK (DB)
- 추정값 표시: JP, FR (스크레이핑 미완료분)
- 출력: data/design_panel/MA_A8_Workbench_Keytruda_SAMPLE.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

BASE_DIR = Path(__file__).parent.parent
OUT_PATH = BASE_DIR / "data" / "design_panel" / "MA_A8_Workbench_Keytruda_SAMPLE.xlsx"

# ────────────────────────────────────────
# 스타일
# ────────────────────────────────────────
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="1F4E78", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_subheader(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_cell(ws, row, col, align=LEFT):
    cell = ws.cell(row=row, column=col)
    cell.alignment = align
    cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ────────────────────────────────────────
# 데이터 (Keytruda 100mg/4mL 1 vial 기준)
# ────────────────────────────────────────
PROJECT = {
    "project_name": "Keytruda 100mg/4mL 2026 상반기 조정 협상",
    "drug_name_en": "Keytruda (pembrolizumab)",
    "drug_name_kr": "키트루다주",
    "manufacturer": "MSD Korea",
    "atc": "L01FF02",
    "sku": "100mg / 4mL / 1 vial",
    "neg_type": "약가 조정",
    "author": "Joseph Kim",
    "date": "2026-04-16",
    "version": "v1.0-SAMPLE",
    "selected_scenario": "B안 (IT 제외)",
    # proposed_ceiling_krw 는 시나리오 계산 후 동적으로 채움
}

# 국가별 현황 (국내 기준 100mg/4mL 1 vial 로 정규화)
COUNTRIES = [
    {
        "country": "JP",
        "source": "MHLW 薬価基準収載品目リスト",
        "url": "https://www.mhlw.go.jp/topics/2024/04/tp20240401-01.html",
        "searched_at": "2026-04-15 14:22",
        "product_match": "キイトルーダ点滴静注100mg 4mL",
        "dosage_form": "주사제",
        "strength": "100mg/4mL",
        "pack": "1 vial",
        "matching_grade": "🟢 Exact",
        "local_price": 279_253.0,
        "currency": "JPY",
        "fx_rate": 9.21,
        "factory_ratio": 0.79,
        "factory_ratio_label": "HIRA 기준 0.79",
        "vat_rate": 0.10,
        "margin": 0.09,
        "note": "※ 추정 샘플 — JP 스크레이퍼 미완",
        "estimate": True,
    },
    {
        "country": "IT",
        "source": "AIFA Lista Classe H (Prezzo Ex-factory)",
        "url": "https://www.aifa.gov.it/en/liste-farmaci-a-h",
        "searched_at": "2026-04-16 10:52",
        "product_match": "KEYTRUDA 1 flaconcino infus EV 4mL 25mg/mL",
        "dosage_form": "주사제",
        "strength": "25mg/mL × 4mL = 100mg",
        "pack": "1 flaconcino",
        "matching_grade": "🟡 Strength-eq",
        "local_price": 3428.0,
        "currency": "EUR",
        "fx_rate": 1526.27,
        "factory_ratio": 1.00,
        "factory_ratio_label": "Ex-factory (직접)",
        "vat_rate": 0.10,
        "margin": 0.09,
        "note": "AIFA 공장도가 직접 공시 — 비율 적용 X",
        "estimate": False,
    },
    {
        "country": "FR",
        "source": "Vidal.fr (Prix public hôpital)",
        "url": "https://www.vidal.fr/medicaments/keytruda-25-mg-ml-sol-diluer-p-perf-173141.html",
        "searched_at": "2026-04-15 14:30",
        "product_match": "KEYTRUDA 25mg/mL sol diluer p perf 4mL",
        "dosage_form": "주사제",
        "strength": "25mg/mL × 4mL = 100mg",
        "pack": "1 flacon",
        "matching_grade": "🟡 Strength-eq",
        "local_price": 2860.0,
        "currency": "EUR",
        "fx_rate": 1526.27,
        "factory_ratio": 0.65,
        "factory_ratio_label": "Vidal 0.65",
        "vat_rate": 0.021,
        "margin": 0.09,
        "note": "※ 추정 샘플 — FR 스크레이퍼 미완",
        "estimate": True,
    },
    {
        "country": "CH",
        "source": "compendium.ch (Publikumspreis, SL)",
        "url": "https://www.compendium.ch/product/1346803/keytruda",
        "searched_at": "2026-04-16 07:32",
        "product_match": "KEYTRUDA Inf Konz 100mg/4mL",
        "dosage_form": "주사제",
        "strength": "100mg/4mL",
        "pack": "1 vial",
        "matching_grade": "🟢 Exact",
        "local_price": 4294.1,
        "currency": "CHF",
        "fx_rate": 1612.87,
        "factory_ratio": 0.65,
        "factory_ratio_label": "Compendium 0.65",
        "vat_rate": 0.077,
        "margin": 0.09,
        "note": "Publikumspreis (공시가)",
        "estimate": False,
    },
    {
        "country": "UK",
        "source": "MIMS online (UK public price)",
        "url": "https://www.mims.co.uk/drugs/cancer/antineoplastics/keytruda",
        "searched_at": "2026-04-16 10:52",
        "product_match": "Keytruda 100mg/4mL conc for soln for inf",
        "dosage_form": "주사제",
        "strength": "100mg/4mL",
        "pack": "1 vial",
        "matching_grade": "🟢 Exact",
        "local_price": 5260.0,
        "currency": "GBP",
        "fx_rate": 1781.91,
        "factory_ratio": 0.73,
        "factory_ratio_label": "HIRA 기준 0.73",
        "vat_rate": 0.00,
        "margin": 0.09,
        "note": "UK 의약품 VAT 면제",
        "estimate": False,
    },
    {
        "country": "DE",
        "source": "Rote Liste (AVP, DocCheck)",
        "url": "https://www.rote-liste.de/rle/detail/26358-0/KEYTRUDA",
        "searched_at": "2026-04-16 07:32",
        "product_match": "KEYTRUDA 25mg/mL Konzentrat",
        "dosage_form": "주사제",
        "strength": "25mg/mL × 4mL = 100mg",
        "pack": "1 vial",
        "matching_grade": "🟡 Strength-eq",
        "local_price": 1868.20,  # 25mg/mL AVP × 4mL = 환산
        "currency": "EUR",
        "fx_rate": 1526.27,
        "factory_ratio": 0.74,
        "factory_ratio_label": "독일 특수공식",
        "vat_rate": 0.19,
        "margin": 0.09,
        "note": "25mg/mL AVP 환산 (×4)",
        "estimate": False,
    },
]


def calc_row(c: dict) -> dict:
    """국가 row 에서 전체 조정가 breakdown 계산."""
    krw_converted = c["local_price"] * c["fx_rate"]
    factory_local = c["local_price"] * c["factory_ratio"]
    factory_krw = krw_converted * c["factory_ratio"]
    vat_applied = factory_krw * (1 + c["vat_rate"])
    adjusted = vat_applied * (1 + c["margin"])
    return {
        **c,
        "krw_converted": int(krw_converted),
        "factory_local": round(factory_local, 2),
        "factory_krw": int(factory_krw),
        "vat_applied_krw": int(vat_applied),
        "adjusted_krw": int(adjusted),
    }


ROWS = [calc_row(c) for c in COUNTRIES]

# ────────────────────────────────────────
# 시나리오 정의
# ────────────────────────────────────────
def build_scenarios(rows):
    def stats(included):
        prices = [r["adjusted_krw"] for r in included]
        if not prices:
            return {}
        avg = sum(prices) // len(prices)
        mn = min(prices)
        mn_country = next(r["country"] for r in included if r["adjusted_krw"] == mn)
        return {
            "countries": " · ".join(r["country"] for r in included),
            "avg": avg,
            "min": mn,
            "min_country": mn_country,
            "min_90": int(mn * 0.9),
            "min_80": int(mn * 0.8),
            "min_70": int(mn * 0.7),
        }

    s_a = stats(rows)  # 전체 6개국
    s_b = stats([r for r in rows if r["country"] != "IT"])  # IT 제외
    s_c = stats(rows)  # 환율 변경 시나리오 — 데모용
    # C안은 가상의 24개월 환율 가정 (+1% KRW 가치 하락)
    s_c = {k: (int(v * 1.01) if isinstance(v, int) else v) for k, v in s_c.items()}

    return {
        "A": {
            "name": "A안 (전체 6개국)",
            "ref_countries": s_a["countries"],
            "fx_basis": "36개월 평균 (KEB하나은행)",
            **s_a,
            "rationale": "전체 A8 참조 — 표준 접근",
            "selected": False,
        },
        "B": {
            "name": "B안 (IT 제외)",
            "ref_countries": s_b["countries"],
            "fx_basis": "36개월 평균 (KEB하나은행)",
            **s_b,
            "rationale": "IT 가격 변동성 제거, 안정 국가 기준",
            "selected": True,
        },
        "C": {
            "name": "C안 (24개월 환율)",
            "ref_countries": s_c["countries"],
            "fx_basis": "24개월 평균 (최근 환율 반영)",
            **s_c,
            "rationale": "최근 환율 반영 — 보수적 기준",
            "selected": False,
        },
    }


SCEN = build_scenarios(ROWS)


# ────────────────────────────────────────
# Sheet 생성
# ────────────────────────────────────────
def create_cover(wb):
    ws = wb.create_sheet("1. Cover", 0)
    set_widths(ws, [28, 60])

    ws["A1"] = "MA Negotiation Workbench"
    ws["A1"].font = Font(bold=True, size=18, color="1F4E78")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    ws["A2"] = "심평원 약가 협상 근거 자료"
    ws["A2"].font = Font(italic=True, size=11, color="666666")
    ws.merge_cells("A2:B2")

    rows = [
        ("프로젝트명",          PROJECT["project_name"]),
        ("대상 제품 (영문)",    PROJECT["drug_name_en"]),
        ("대상 제품 (국문)",    PROJECT["drug_name_kr"]),
        ("제조사",              PROJECT["manufacturer"]),
        ("ATC 코드",            PROJECT["atc"]),
        ("국내 기준 SKU",       PROJECT["sku"]),
        ("협상 유형",           PROJECT["neg_type"]),
        ("",                    ""),
        ("선정 시나리오",       PROJECT["selected_scenario"]),
        ("제안 상한가 (KRW)",   f"₩ {SCEN['B']['min_90']:,}"),
        ("산정 기준",           f"A8 최저 × 90%  (최저: {SCEN['B']['min_country']} {SCEN['B']['min']:,} KRW)"),
        ("",                    ""),
        ("작성자",              PROJECT["author"]),
        ("작성일",              PROJECT["date"]),
        ("버전",                PROJECT["version"]),
        ("데이터 최신성",       "각국 소스 스크레이핑 시점 별도 기록 (Sheet 4 참조)"),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, color="1F4E78")
        ws.cell(row=i, column=1).alignment = LEFT
        ws.cell(row=i, column=2, value=v).alignment = LEFT
        if k == "제안 상한가 (KRW)":
            ws.cell(row=i, column=2).font = Font(bold=True, size=14, color="C00000")

    # footer note
    ws.cell(row=len(rows) + 6, column=1, value="※ 본 문서는 MA Negotiation Workbench v1.0 템플릿 샘플입니다.")
    ws.cell(row=len(rows) + 6, column=1).font = Font(italic=True, color="888888", size=9)


def create_a8_summary(wb):
    ws = wb.create_sheet("2. A8 Summary")
    set_widths(ws, [8, 35, 10, 14, 11, 16, 16, 16, 10])

    headers = ["국가", "원통화 제품", "통화", "현지약가", "환율", "KRW 환산", "공장도가(KRW)", "조정가(KRW)", "포함"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, range(1, len(headers) + 1))

    for idx, r in enumerate(ROWS, start=2):
        ws.cell(row=idx, column=1, value=r["country"])
        ws.cell(row=idx, column=2, value=r["product_match"] + (" ※추정" if r["estimate"] else ""))
        ws.cell(row=idx, column=3, value=r["currency"])
        ws.cell(row=idx, column=4, value=r["local_price"])
        ws.cell(row=idx, column=5, value=r["fx_rate"])
        ws.cell(row=idx, column=6, value=r["krw_converted"])
        ws.cell(row=idx, column=7, value=r["factory_krw"])
        ws.cell(row=idx, column=8, value=r["adjusted_krw"])
        ws.cell(row=idx, column=9, value="✅" if r["country"] != "IT" else "❌")  # B안 기준
        for c in range(1, len(headers) + 1):
            style_cell(ws, idx, c, align=RIGHT if c in (4, 5, 6, 7, 8) else CENTER)

    # 통계 섹션
    start = len(ROWS) + 3
    ws.cell(row=start, column=1, value="A8 통계 (B안 — IT 제외)").font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=9)

    stats = SCEN["B"]
    stat_rows = [
        ("A8 평균",        f"{stats['avg']:,}"),
        ("A8 최저",        f"{stats['min']:,}  ({stats['min_country']})"),
        ("최저 × 90%",     f"{stats['min_90']:,}"),
        ("최저 × 80%",     f"{stats['min_80']:,}"),
        ("최저 × 70%",     f"{stats['min_70']:,}"),
    ]
    for i, (k, v) in enumerate(stat_rows, start=start + 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

    # 막대 차트
    chart = BarChart()
    chart.type = "bar"
    chart.title = "국가별 조정가 (KRW)"
    chart.y_axis.title = "국가"
    chart.x_axis.title = "조정가 (KRW)"
    data = Reference(ws, min_col=8, min_row=1, max_row=len(ROWS) + 1, max_col=8)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(ROWS) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, f"A{start + len(stat_rows) + 2}")


def create_adjustment_logic(wb):
    ws = wb.create_sheet("3. Adjustment Logic")
    set_widths(ws, [8, 32, 12, 10, 14, 11, 14, 14, 7, 14, 8, 14])

    headers = ["국가", "자료원", "현지약가", "환율(36mo)", "KRW 환산",
               "공장도비율", "공장도가(현지)", "공장도가(KRW)",
               "VAT", "VAT 적용", "마진", "조정가(KRW)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, range(1, len(headers) + 1))

    for idx, r in enumerate(ROWS, start=2):
        cells = [
            r["country"], r["source"], r["local_price"], r["fx_rate"], r["krw_converted"],
            r["factory_ratio_label"], r["factory_local"], r["factory_krw"],
            f"{r['vat_rate']*100:.1f}%", r["vat_applied_krw"],
            f"{r['margin']*100:.0f}%", r["adjusted_krw"],
        ]
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=idx, column=c, value=v)
            cell.border = BORDER
            cell.alignment = RIGHT if c in (3, 4, 5, 7, 8, 10, 12) else CENTER

    # 공식 설명
    note_row = len(ROWS) + 3
    ws.cell(row=note_row, column=1, value="계산 공식").font = Font(bold=True, color="1F4E78")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=12)

    formulas = [
        "현지약가 × 환율 = KRW 환산",
        "KRW 환산 × 공장도비율 = 공장도가(KRW)    (IT: Ex-factory 직접 사용)",
        "공장도가(KRW) × (1 + VAT) = VAT 적용가",
        "VAT 적용가 × (1 + 유통마진) = 조정가(KRW)",
        "환율: KEB하나은행 36개월 평균 (기본값, Assumptions 시트에서 수정 가능)",
        "공장도비율: HIRA 고시값 (국가별 상이 — JP 0.79 / IT 1.00 / FR 0.65 / CH 0.65 / UK 0.73 / DE 특수)",
    ]
    for i, f in enumerate(formulas, start=note_row + 1):
        ws.cell(row=i, column=1, value=f"• {f}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)


def create_source_raw(wb):
    ws = wb.create_sheet("4. Source Raw")
    set_widths(ws, [8, 32, 55, 18, 38, 12, 10])

    headers = ["국가", "사이트", "URL", "조회일시", "매칭 제품명", "원본 가격", "통화"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, range(1, len(headers) + 1))

    for idx, r in enumerate(ROWS, start=2):
        cells = [r["country"], r["source"], r["url"], r["searched_at"], r["product_match"], r["local_price"], r["currency"]]
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=idx, column=c, value=v)
            cell.border = BORDER
            cell.alignment = LEFT
            if c == 3:
                cell.hyperlink = r["url"]
                cell.font = Font(color="0563C1", underline="single")


def create_product_matching(wb):
    ws = wb.create_sheet("5. Product Matching")
    set_widths(ws, [8, 32, 38, 10, 22, 14, 18, 30])

    headers = ["국가", "소스", "추출 제품명", "제형", "강도", "Pack", "국내 매칭", "일관성 등급"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, range(1, len(headers) + 1))

    for idx, r in enumerate(ROWS, start=2):
        cells = [r["country"], r["source"], r["product_match"],
                 r["dosage_form"], r["strength"], r["pack"],
                 PROJECT["sku"], r["matching_grade"]]
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=idx, column=c, value=v)
            cell.border = BORDER
            cell.alignment = LEFT

    # 범례
    note_row = len(ROWS) + 3
    notes = [
        "일관성 등급 — 데이터 일관성 내부 검증용 (심평원 심사 영역 아님)",
        "🟢 Exact          : 국내 기준 SKU 와 동일 (100mg/4mL/vial)",
        "🟡 Strength-eq    : 강도/포장 환산 필요 (예: 25mg/mL × 4mL = 100mg)",
        "🔴 Mismatch       : 매칭 실패 — 사용자 수동 확인 필요",
    ]
    for i, n in enumerate(notes, start=note_row):
        c = ws.cell(row=i, column=1, value=n)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        if i == note_row:
            c.font = Font(bold=True, color="1F4E78")


def create_hta_matrix(wb):
    ws = wb.create_sheet("6. HTA Matrix")
    set_widths(ws, [38, 10, 10, 10, 10, 10, 10, 40])

    ws["A1"] = "HTA Matrix — Phase 2 에서 채움"
    ws["A1"].font = Font(bold=True, size=14, color="C00000")
    ws["A2"] = "Phase 1 MVP 에서는 빈 시트로 유지. Phase 2 진입 시 적응증별 국가 매트릭스 채움."
    ws["A2"].font = Font(italic=True, color="666666")

    headers = ["적응증", "FDA", "EMA", "PBAC", "CADTH", "NICE", "SMC", "A8 가격형성 영향"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, range(1, len(headers) + 1))

    # 샘플 2줄 (회색)
    sample = [
        ["1L NSCLC (단독)", "✅", "✅", "✅", "✅", "✅", "✅", "전 국가 공통 — 기본 근거 적응증"],
        ["1L NSCLC (Pembro+Chemo)", "✅", "✅", "✅", "❌", "✅", "⚠️", "CADTH 미급여 → A8 내 가격 낮음"],
    ]
    for idx, row in enumerate(sample, start=5):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=idx, column=c, value=v)
            cell.border = BORDER
            cell.alignment = CENTER
            cell.font = Font(color="888888", italic=True)


def create_assumptions(wb):
    ws = wb.create_sheet("7. Assumptions")
    set_widths(ws, [30, 18, 40, 30])

    ws["A1"] = "가정치 명세 — 대시보드에서 편집 가능"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E78")
    ws.merge_cells("A1:D1")

    ws["A2"] = "HIRA 고시값이 기본값. 변경 시 Audit Log 에 기록됨."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:D2")

    headers = ["항목", "값", "출처", "비고"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, range(1, len(headers) + 1))

    assumptions = [
        ("[환율 — KEB하나은행 36개월 평균]", "", "", ""),
        ("JPY → KRW",             9.21,    "기준기간 2023-04 ~ 2026-03",   ""),
        ("EUR → KRW",             1526.27, "기준기간 2023-04 ~ 2026-03",   ""),
        ("CHF → KRW",             1612.87, "기준기간 2023-04 ~ 2026-03",   ""),
        ("GBP → KRW",             1781.91, "기준기간 2023-04 ~ 2026-03",   ""),
        ("",                      "",       "",                            ""),
        ("[공장도비율 — HIRA 고시]", "",     "",                            ""),
        ("JP",                    0.79,    "HIRA 고시",                   ""),
        ("IT",                    1.00,    "AIFA Ex-factory 직접",        "특수: 비율 적용 X"),
        ("FR",                    0.65,    "Vidal 기준",                  ""),
        ("CH",                    0.65,    "Compendium 기준",             ""),
        ("UK",                    0.73,    "HIRA 고시",                   ""),
        ("DE",                    0.74,    "독일 특수공식 (HIRA 고시)",    ""),
        ("",                      "",       "",                            ""),
        ("[VAT — 각국 세법]",      "",     "",                            ""),
        ("JP",                    "10.0%", "소비세",                      ""),
        ("IT",                    "10.0%", "의약품 경감세율",              ""),
        ("FR",                    "2.1%",  "의약품 경감세율",              ""),
        ("CH",                    "7.7%",  "표준세율",                    ""),
        ("UK",                    "0.0%",  "의약품 VAT 면제",              ""),
        ("DE",                    "19.0%", "표준세율",                    ""),
        ("",                      "",       "",                            ""),
        ("[유통마진]",             "",     "",                            ""),
        ("HIRA 표준 유통거래폭",   "9.0%",  "HIRA 고시",                   "전 국가 공통 적용"),
        ("",                      "",       "",                            ""),
        ("[제외 국가 — 시나리오 B 한정]", "", "",                           ""),
        ("IT",                    "제외",  "가격 변동성",                 "최근 6개월 AIFA 재분류"),
    ]
    for i, (k, v, src, note) in enumerate(assumptions, start=5):
        is_section = k.startswith("[")
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, color="1F4E78") if is_section else Font()
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=src)
        ws.cell(row=i, column=4, value=note)
        for c in range(1, 5):
            ws.cell(row=i, column=c).alignment = LEFT
            if not is_section and v != "":
                ws.cell(row=i, column=c).border = BORDER


def create_scenarios(wb):
    ws = wb.create_sheet("8. Scenarios")
    set_widths(ws, [22, 25, 25, 25])

    ws["A1"] = "시나리오 비교"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:D1")

    headers = ["항목", SCEN["A"]["name"], SCEN["B"]["name"], SCEN["C"]["name"]]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, range(1, len(headers) + 1))

    rows = [
        ("참조 국가",      "ref_countries"),
        ("환율 기준",      "fx_basis"),
        ("A8 평균 (KRW)",  "avg"),
        ("A8 최저 (KRW)",  "min"),
        ("최저 × 90%",     "min_90"),
        ("최저 × 80%",     "min_80"),
        ("최저 × 70%",     "min_70"),
        ("전략 근거",      "rationale"),
    ]
    for i, (label, key) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        for j, s in enumerate(["A", "B", "C"], start=2):
            val = SCEN[s][key]
            if isinstance(val, int) and key != "fx_basis":
                val = f"{val:,}"
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = LEFT
            cell.border = BORDER

    final_row = 4 + len(rows) + 1
    ws.cell(row=final_row, column=1, value="최종 선택").font = Font(bold=True, color="C00000")
    for j, s in enumerate(["A", "B", "C"], start=2):
        mark = "✅ 선정" if SCEN[s]["selected"] else ""
        cell = ws.cell(row=final_row, column=j, value=mark)
        cell.alignment = CENTER
        cell.font = Font(bold=True, color="008000" if mark else "000000", size=13)
        cell.border = BORDER


def create_audit_log(wb):
    ws = wb.create_sheet("9. Audit Log")
    set_widths(ws, [18, 14, 18, 24, 16, 16, 40])

    headers = ["타임스탬프", "사용자", "변경 시트", "항목", "이전 값", "신규 값", "사유"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, range(1, len(headers) + 1))

    log_entries = [
        ("2026-04-16 09:12", "Joseph Kim", "4. Source Raw",     "Keytruda 전 국가 스크레이핑 실행", "-", "6개국 완료(JP/FR 추정)", "Workbench 초기 생성"),
        ("2026-04-16 10:15", "Joseph Kim", "7. Assumptions",    "IT 포함여부",                    "✅",        "❌",         "가격 변동성 (시나리오 B)"),
        ("2026-04-16 10:22", "Joseph Kim", "8. Scenarios",      "최종 선택",                      "A안",       "B안",        "내부 협상회의 결정"),
        ("2026-04-16 10:30", "Joseph Kim", "1. Cover",          "제안 상한가",                    "자동값",    "5,769,948",  "B안 최저 × 90% 확정"),
    ]
    for idx, entry in enumerate(log_entries, start=2):
        for c, v in enumerate(entry, start=1):
            cell = ws.cell(row=idx, column=c, value=v)
            cell.border = BORDER
            cell.alignment = LEFT


# ────────────────────────────────────────
# main
# ────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    create_cover(wb)
    create_a8_summary(wb)
    create_adjustment_logic(wb)
    create_source_raw(wb)
    create_product_matching(wb)
    create_hta_matrix(wb)
    create_assumptions(wb)
    create_scenarios(wb)
    create_audit_log(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)

    # B안 제안 상한가 검증 출력
    b = SCEN["B"]
    print(f"✅ 샘플 워크북 생성: {OUT_PATH}")
    print(f"   B안 A8 최저 : {b['min']:,} KRW ({b['min_country']})")
    print(f"   B안 × 90%   : {b['min_90']:,} KRW  ← 제안 상한가")
    print(f"\n국가별 조정가:")
    for r in ROWS:
        flag = " (추정)" if r["estimate"] else ""
        print(f"   {r['country']}: {r['adjusted_krw']:>12,} KRW  [{r['matching_grade']}]{flag}")


if __name__ == "__main__":
    main()
