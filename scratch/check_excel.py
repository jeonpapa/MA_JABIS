from pathlib import Path
import openpyxl

excel_path = Path("/Users/kimjeong-ae/MA_AI_Dossier/data/hira_pipeline/hira_committee_master.xlsx")
print("Exists:", excel_path.exists())
if excel_path.exists():
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print("Sheets:", wb.sheetnames)
    ws = wb.active
    # Count rows
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
    print("Total rows (including header):", row_count)
